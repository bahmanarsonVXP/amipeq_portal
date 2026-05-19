#!/usr/bin/env python3
"""
Phase 3 - Create Notes for each imported Opportunity
"""

import json
import time
import urllib.request
import urllib.error
import ssl
import os
import openpyxl
import re
from datetime import datetime, timedelta

env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '.env')
if os.path.exists(env_path):
    with open(env_path) as f:
        for line in f:
            if line.strip() and not line.strip().startswith('#'):
                key, val = line.strip().split('=', 1)
                os.environ[key.strip()] = val.strip()

BASE_URL = os.environ.get('TWENTY_BASE_URL', 'https://twenty-production-7352.up.railway.app')
API_KEY = os.environ.get('TWENTY_API_KEY', '')
TEST_LIMIT = 50
EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'SUIVIS CLIENTS 2026.xlsx')
MAPPINGS_FILE = os.path.join(os.path.dirname(__file__), 'mappings.json')

ssl_ctx = ssl.create_default_context()

def api_request(method, path, data=None):
    url = f"{BASE_URL}{path}"
    headers = {
        'Authorization': f'Bearer {API_KEY}',
        'Content-Type': 'application/json',
    }
    body = json.dumps(data).encode('utf-8') if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, context=ssl_ctx) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        error_body = e.read().decode('utf-8')
        try:
            return {'error': json.loads(error_body), 'status': e.code}
        except:
            return {'error': error_body, 'status': e.code}

def safe_str(val):
    if val is None:
        return ''
    return str(val).strip()

def get_cell_color(cell):
    if cell.fill and cell.fill.fgColor:
        rgb = str(cell.fill.fgColor.rgb) if cell.fill.fgColor.rgb else ''
        if rgb == 'FF92D050' or rgb == '0092D050':
            return 'Gagné'
        elif rgb in ('FFA5A5A5', 'FFBFBFBF', '00A5A5A5', '00BFBFBF'):
            return 'Refusé'
    return 'En attente'

def main():
    print("=== PHASE 3 : CRÉATION DES NOTES ===\n")

    # Load mappings
    with open(MAPPINGS_FILE, 'r') as f:
        mappings = json.load(f)

    # Read Excel again to get context for notes
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheets = ['2023', '2024', '2025', '2026']

    all_rows = []
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        year = int(sheet_name)
        for row_idx in range(2, ws.max_row + 1):
            num_ste = safe_str(ws.cell(row=row_idx, column=3).value)
            client_name = safe_str(ws.cell(row=row_idx, column=4).value)
            num_devis = safe_str(ws.cell(row=row_idx, column=8).value)
            norme = safe_str(ws.cell(row=row_idx, column=12).value)
            statut = get_cell_color(ws.cell(row=row_idx, column=1))

            if not num_ste or not client_name:
                continue

            all_rows.append({
                'year': year,
                'num_ste': num_ste,
                'client_name': client_name,
                'num_devis': num_devis,
                'norme': norme,
                'statut': statut,
            })

    all_rows = all_rows[:TEST_LIMIT]

    notes_created = 0
    notes_errors = 0
    notes_skipped = 0

    for row in all_rows:
        opp_key = row['num_devis'] or f"{row['num_ste']}-{row['year']}"
        opp_id = mappings['opportunities'].get(opp_key)
        company_id = mappings['companies'].get(row['num_ste'])

        if not opp_id:
            notes_skipped += 1
            continue

        # Build note
        note_body = (
            f"Importé depuis SUIVIS_CLIENTS_2026.xlsx\n"
            f"Année: {row['year']}\n"
            f"Client: {row['client_name']}\n"
            f"Norme originale: {row['norme']}\n"
            f"Statut: {row['statut']}"
        )

        note_data = {
            'title': f"Import Excel - Devis {opp_key}",
        }

        result = api_request('POST', '/rest/notes', note_data)
        time.sleep(0.7)

        if 'error' in result:
            print(f"  ERR: {opp_key} → {json.dumps(result['error'])[:200]}")
            notes_errors += 1
        else:
            note_id = result.get('data', {}).get('createNote', {}).get('id') or result.get('id')
            if not note_id and 'data' in result:
                for k in result['data']:
                    if isinstance(result['data'][k], dict) and 'id' in result['data'][k]:
                        note_id = result['data'][k]['id']
                        break

            if note_id:
                print(f"  OK: Note {opp_key} → {note_id}")
                notes_created += 1

                # Create noteTarget to link to company
                if company_id:
                    target_data = {
                        'noteId': note_id,
                        'targetCompanyId': company_id,
                    }
                    target_result = api_request('POST', '/rest/noteTargets', target_data)
                    time.sleep(0.7)
                    if 'error' in target_result:
                        print(f"    ERR noteTarget: {json.dumps(target_result['error'])[:150]}")
                    else:
                        print(f"    OK noteTarget linked to company")
            else:
                print(f"  ???: {opp_key} → {json.dumps(result)[:300]}")
                notes_errors += 1

    print(f"\n=== RÉSUMÉ PHASE 3 ===")
    print(f"  Notes créées: {notes_created}")
    print(f"  Erreurs: {notes_errors}")
    print(f"  Ignorées: {notes_skipped}")

if __name__ == '__main__':
    main()
