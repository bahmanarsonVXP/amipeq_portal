#!/usr/bin/env python3
"""
Phase 2 - Import data from Excel to Twenty CRM (TEST MODE: 50 lines)
Reads SUIVIS CLIENTS 2026.xlsx, deduplicates, imports Companies → Persons → Opportunities
"""

import openpyxl
import json
import re
import time
import urllib.request
import urllib.error
import ssl
import os
from datetime import datetime, timedelta

# Config
env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '.env')
if os.path.exists(env_path):
    with open(env_path) as f:
        for line in f:
            if line.strip() and not line.strip().startswith('#'):
                key, val = line.strip().split('=', 1)
                os.environ[key.strip()] = val.strip()

BASE_URL = os.environ.get('TWENTY_BASE_URL', 'https://twenty-production-7352.up.railway.app')
API_KEY = os.environ.get('TWENTY_API_KEY', '')
TEST_LIMIT = 50  # Mode test: 50 lignes max
EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'SUIVIS CLIENTS 2026.xlsx')
MAPPINGS_FILE = os.path.join(os.path.dirname(__file__), 'mappings.json')
LOG_FILE = os.path.join(os.path.dirname(__file__), 'import_log.json')

SHEETS_TO_IMPORT = ['2023', '2024', '2025', '2026']

ssl_ctx = ssl.create_default_context()

# ============================================================
# API HELPERS
# ============================================================

def api_request(method, path, data=None):
    """Make HTTP request to Twenty CRM API"""
    url = f"{BASE_URL}{path}"
    headers = {
        'Authorization': f'Bearer {API_KEY}',
        'Content-Type': 'application/json',
    }
    body = json.dumps(data).encode('utf-8') if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, context=ssl_ctx) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        error_body = e.read().decode('utf-8')
        try:
            return {'error': json.loads(error_body), 'status': e.code}
        except:
            return {'error': error_body, 'status': e.code}

def gql_request(endpoint, query, variables=None):
    """Make GraphQL request"""
    payload = {'query': query}
    if variables:
        payload['variables'] = variables
    return api_request('POST', endpoint, payload)

def rest_create(plural_name, data):
    """Create a record via REST API"""
    return api_request('POST', f'/rest/{plural_name}', data)

# ============================================================
# EXCEL READING
# ============================================================

def excel_date_to_iso(value):
    """Convert Excel serial date to ISO date string"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, (int, float)):
        try:
            dt = datetime(1899, 12, 30) + timedelta(days=int(value))
            return dt.strftime('%Y-%m-%d')
        except:
            return None
    if isinstance(value, str):
        # Try parsing as date string
        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%m/%Y']:
            try:
                return datetime.strptime(value, fmt).strftime('%Y-%m-%d')
            except:
                continue
    return None

def get_cell_color(cell):
    """Get the fill color of a cell to determine devis status"""
    if cell.fill and cell.fill.fgColor:
        rgb = str(cell.fill.fgColor.rgb) if cell.fill.fgColor.rgb else ''
        if rgb == 'FF92D050' or rgb == '0092D050':
            return 'GAGNE'
        elif rgb in ('FFA5A5A5', 'FFBFBFBF', '00A5A5A5', '00BFBFBF'):
            return 'REFUSE'
    return 'EN_ATTENTE'

def extract_departement(cp_raw):
    """Extract department number from postal code"""
    if cp_raw is None:
        return None, None
    try:
        cp = str(int(float(cp_raw))).zfill(5)
    except (ValueError, TypeError):
        return None, None

    if cp[:3] in ('971', '972', '973', '974'):
        num = cp[:3]
    elif cp[:2] == '20':
        num = '2A' if int(cp) < 20200 else '2B'
    else:
        num = cp[:2]

    # Build SELECT value
    dept_map = {
        '01':'Ain','02':'Aisne','03':'Allier','04':'Alpes_de_Haute_Provence','05':'Hautes_Alpes',
        '06':'Alpes_Maritimes','07':'Ardeche','08':'Ardennes','09':'Ariege','10':'Aube',
        '11':'Aude','12':'Aveyron','13':'Bouches_du_Rhone','14':'Calvados','15':'Cantal',
        '16':'Charente','17':'Charente_Maritime','18':'Cher','19':'Correze',
        '2A':'Corse_du_Sud','2B':'Haute_Corse',
        '21':'Cote_d_Or','22':'Cotes_d_Armor','23':'Creuse',
        '24':'Dordogne','25':'Doubs','26':'Drome','27':'Eure','28':'Eure_et_Loir','29':'Finistere',
        '30':'Gard','31':'Haute_Garonne','32':'Gers','33':'Gironde','34':'Herault','35':'Ille_et_Vilaine',
        '36':'Indre','37':'Indre_et_Loire','38':'Isere','39':'Jura','40':'Landes','41':'Loir_et_Cher',
        '42':'Loire','43':'Haute_Loire','44':'Loire_Atlantique','45':'Loiret','46':'Lot','47':'Lot_et_Garonne',
        '48':'Lozere','49':'Maine_et_Loire','50':'Manche','51':'Marne','52':'Haute_Marne','53':'Mayenne',
        '54':'Meurthe_et_Moselle','55':'Meuse','56':'Morbihan','57':'Moselle','58':'Nievre','59':'Nord',
        '60':'Oise','61':'Orne','62':'Pas_de_Calais','63':'Puy_de_Dome','64':'Pyrenees_Atlantiques','65':'Hautes_Pyrenees',
        '66':'Pyrenees_Orientales','67':'Bas_Rhin','68':'Haut_Rhin','69':'Rhone','70':'Haute_Saone','71':'Saone_et_Loire',
        '72':'Sarthe','73':'Savoie','74':'Haute_Savoie','75':'Paris','76':'Seine_Maritime','77':'Seine_et_Marne',
        '78':'Yvelines','79':'Deux_Sevres','80':'Somme','81':'Tarn','82':'Tarn_et_Garonne','83':'Var',
        '84':'Vaucluse','85':'Vendee','86':'Vienne','87':'Haute_Vienne','88':'Vosges','89':'Yonne',
        '90':'Territoire_de_Belfort','91':'Essonne','92':'Hauts_de_Seine','93':'Seine_Saint_Denis','94':'Val_de_Marne','95':'Val_d_Oise',
        '971':'Guadeloupe','972':'Martinique','973':'Guyane','974':'La_Reunion'
    }
    name = dept_map.get(num, 'INCONNU')
    select_value = f"DEPT_{num}_{name.upper()}"
    return num, select_value

def classify_client(name):
    """Classify client type based on name keywords"""
    upper = name.upper()
    if any(kw in upper for kw in ['COLLEGE', 'COLLÈGE', 'LYCEE', 'LYCÉE', 'ECOLE', 'ÉCOLE', 'GROUPE SCOLAIRE', 'SCOLAIRE', 'INSTITUTION']):
        type_client = 'ETABLISSEMENT_SCOLAIRE'
        if 'COLLEGE' in upper or 'COLLÈGE' in upper:
            sous_type = 'COLLEGE'
        elif 'LYCEE' in upper or 'LYCÉE' in upper:
            sous_type = 'LYCEE'
        elif 'ECOLE' in upper or 'ÉCOLE' in upper:
            sous_type = 'ECOLE'
        else:
            sous_type = None
    elif any(kw in upper for kw in ['MAIRIE', 'COMMUNAUT', 'AGGLO', 'CC DE', 'CC DU']):
        type_client = 'MAIRIE_COLLECTIVITE'
        if 'MAIRIE' in upper:
            sous_type = 'MAIRIE'
        elif 'COMMUNAUT' in upper or 'AGGLO' in upper:
            sous_type = 'COMMUNAUTE_DE_COMMUNES'
        else:
            sous_type = None
    elif 'EHPAD' in upper:
        type_client = 'MAIRIE_COLLECTIVITE'
        sous_type = 'EHPAD'
    else:
        type_client = 'ENTREPRISE_TPE_PME'
        sous_type = None
    return type_client, sous_type

def parse_norme(raw):
    """Parse NORME column into prestation, nature, modalite"""
    if not raw:
        return ['DUERP'], 'CREATION', None

    s = str(raw).strip()

    # 1. MODALITÉ
    modalite = None
    modal_patterns = [
        (r'(?i)\bà distance\b', 'A_DISTANCE'),
        (r'(?i)\ba distance\b', 'A_DISTANCE'),
        (r'(?i)\bdématérialisé[e]?\b', 'A_DISTANCE'),
        (r'(?i)\bdistanciel\b', 'A_DISTANCE'),
        (r'(?i)\bdistance\b', 'A_DISTANCE'),
        (r'(?i)\bsur site\b', 'SUR_SITE'),
        (r'(?i)\bprésentiel\b', 'SUR_SITE'),
    ]
    for pattern, modal in modal_patterns:
        if re.search(pattern, s):
            modalite = modal
            s = re.sub(pattern, '', s).strip()
            break
    if re.search(r'(?i)sur site ou', s):
        modalite = 'SUR_SITE_OU_A_DISTANCE'
        s = re.sub(r'(?i)sur site ou', '', s).strip()

    # 2. NATURE
    nature = 'CREATION'
    if re.search(r'(?i)\bcontrat\s+maj\b', s):
        nature = 'CONTRAT_MAJ'
        s = re.sub(r'(?i)\bcontrat\s+maj\b', '', s).strip()
    elif re.search(r'(?i)\bmaj\b', s):
        nature = 'MISE_A_JOUR'
        s = re.sub(r'(?i)\bmaj\b', '', s).strip()

    # 3. PRESTATIONS
    s_upper = s.upper()
    for noise in ['+ CLASSEUR', 'CLASSEUR', 'ET/OU', 'SEUL', 'SS ', 'DEVIS ', 'SIGNES', 'SIGNE', 'CLIENT RÉCENT', 'CLIENT']:
        s_upper = s_upper.replace(noise, '')
    s_upper = re.sub(r'[+/,\-]', ' ', s_upper)
    s_upper = ' '.join(s_upper.split())

    prestations = set()
    full_maps = {
        'DOCUMENT UNIQUE': 'DUERP', 'DUERP': 'DUERP', 'DUER': 'DUERP',
        'DUEP': 'DUERP', 'DU': 'DUERP',
        'PPMS': 'PPMS', 'PMMS': 'PPMS',
        'RPS': 'RPS', 'ENTRETIENS INDIVIDUELS': 'RPS', 'ENTRETIENS': 'RPS',
        'PSE': 'PSE', 'PLAN BLANC ET BLEU': 'PSE',
        'COVID': 'COVID', 'COVID 19': 'COVID',
        'RGPD': 'RGPD',
    }
    temp = s_upper
    for expr in sorted(full_maps.keys(), key=len, reverse=True):
        if expr in temp:
            prestations.add(full_maps[expr])
            temp = temp.replace(expr, ' ', 1)
    if not prestations:
        prestations.add('DUERP')

    return sorted(prestations), nature, modalite

def parse_contact_name(contact, civilite):
    """Parse contact name into firstName, lastName"""
    if not contact:
        return None, None

    name = str(contact).strip()
    # Remove common prefixes
    name = re.sub(r'^(M\.|Mme|MME|M |Melle)\s*', '', name).strip()

    parts = name.split()
    if len(parts) == 0:
        return None, None
    elif len(parts) == 1:
        return None, parts[0].upper()
    else:
        # Check if first part is all uppercase (likely LASTNAME)
        if parts[0].isupper() and len(parts) > 1:
            return ' '.join(parts[1:]), parts[0]
        else:
            # Assume "FirstName LASTNAME" or "LASTNAME FirstName"
            # If last part is uppercase, it's the lastname
            if parts[-1].isupper() and not parts[0].isupper():
                return ' '.join(parts[:-1]), parts[-1]
            else:
                return parts[0], ' '.join(parts[1:]).upper()

def safe_str(val):
    """Safely convert value to string"""
    if val is None:
        return ''
    return str(val).strip()

def safe_float(val):
    """Safely convert value to float"""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

# ============================================================
# MAIN
# ============================================================

def main():
    print(f"=== PHASE 2 : IMPORT DONNÉES (MODE TEST - {TEST_LIMIT} lignes) ===\n")

    # Load existing mappings if any
    mappings = {'companies': {}, 'persons': {}, 'opportunities': {}}
    if os.path.exists(MAPPINGS_FILE):
        with open(MAPPINGS_FILE, 'r') as f:
            mappings = json.load(f)

    log = []

    # --------------------------------------------------------
    # STEP 1: Read Excel and collect all rows
    # --------------------------------------------------------
    print("1. Lecture du fichier Excel...")
    wb = openpyxl.load_workbook(EXCEL_FILE)

    all_rows = []
    for sheet_name in SHEETS_TO_IMPORT:
        if sheet_name not in wb.sheetnames:
            print(f"  Onglet {sheet_name} non trouvé, ignoré")
            continue

        ws = wb[sheet_name]
        year = int(sheet_name)
        row_count = 0

        for row_idx in range(2, ws.max_row + 1):
            # Read cell values
            prosp = safe_str(ws.cell(row=row_idx, column=1).value)
            date_raw = ws.cell(row=row_idx, column=2).value
            num_ste = safe_str(ws.cell(row=row_idx, column=3).value)
            client_name = safe_str(ws.cell(row=row_idx, column=4).value)
            civilite = safe_str(ws.cell(row=row_idx, column=5).value)
            contact = safe_str(ws.cell(row=row_idx, column=6).value)
            commercial = safe_str(ws.cell(row=row_idx, column=7).value)
            num_devis = safe_str(ws.cell(row=row_idx, column=8).value)
            date_devis_raw = ws.cell(row=row_idx, column=9).value
            offre1 = safe_float(ws.cell(row=row_idx, column=10).value)
            offre2 = safe_float(ws.cell(row=row_idx, column=11).value)
            norme = safe_str(ws.cell(row=row_idx, column=12).value)
            adresse1 = safe_str(ws.cell(row=row_idx, column=17).value)
            adresse2 = safe_str(ws.cell(row=row_idx, column=18).value)
            cp = ws.cell(row=row_idx, column=19).value
            ville = safe_str(ws.cell(row=row_idx, column=20).value)
            telephone = safe_str(ws.cell(row=row_idx, column=21).value)
            relance_tel = safe_str(ws.cell(row=row_idx, column=22).value)
            email = safe_str(ws.cell(row=row_idx, column=23).value)
            date_docs_raw = ws.cell(row=row_idx, column=24).value

            # Skip empty rows
            if not num_ste or not client_name:
                continue

            # Get cell color for status
            statut = get_cell_color(ws.cell(row=row_idx, column=1))

            all_rows.append({
                'year': year,
                'row_idx': row_idx,
                'prosp': prosp,
                'date': date_raw,
                'num_ste': num_ste,
                'client_name': client_name,
                'civilite': civilite,
                'contact': contact,
                'commercial': commercial,
                'num_devis': num_devis,
                'date_devis': date_devis_raw,
                'offre1': offre1,
                'offre2': offre2,
                'norme': norme,
                'adresse1': adresse1,
                'adresse2': adresse2,
                'cp': cp,
                'ville': ville,
                'telephone': telephone,
                'relance_tel': relance_tel,
                'email': email,
                'date_docs': date_docs_raw,
                'statut': statut,
            })

            row_count += 1

        print(f"  Onglet {sheet_name}: {row_count} lignes lues")

    print(f"  Total: {len(all_rows)} lignes lues")

    # Limit to TEST_LIMIT
    all_rows = all_rows[:TEST_LIMIT]
    print(f"  Mode test: limité à {len(all_rows)} lignes\n")

    # --------------------------------------------------------
    # STEP 2: Deduplicate
    # --------------------------------------------------------
    print("2. Dédoublonnage...")

    # Companies: keep last occurrence per N° Société
    companies_map = {}  # num_ste -> company data
    company_devis_status = {}  # num_ste -> list of (year, statut)

    for row in all_rows:
        num_ste = row['num_ste']
        if num_ste not in companies_map:
            companies_map[num_ste] = row
            company_devis_status[num_ste] = []
        else:
            # Keep the most recent data
            if row['year'] >= companies_map[num_ste]['year']:
                companies_map[num_ste] = row
        company_devis_status[num_ste].append((row['year'], row['statut']))

    # Persons: deduplicate by (num_ste, contact_name)
    persons_map = {}  # "num_ste|contact" -> person data
    for row in all_rows:
        if not row['contact']:
            continue
        key = f"{row['num_ste']}|{row['contact']}"
        if key not in persons_map:
            persons_map[key] = row

    print(f"  Companies uniques: {len(companies_map)}")
    print(f"  Persons uniques: {len(persons_map)}")
    print(f"  Opportunities: {len(all_rows)}\n")

    # --------------------------------------------------------
    # STEP 3: Import Companies
    # --------------------------------------------------------
    print("3. Import Companies...")
    companies_created = 0
    companies_errors = 0

    for num_ste, row in companies_map.items():
        if num_ste in mappings['companies']:
            print(f"  SKIP: {row['client_name']} (déjà importé)")
            continue

        dept_num, dept_select = extract_departement(row['cp'])
        type_client, sous_type = classify_client(row['client_name'])

        # Determine statut
        statuts = company_devis_status.get(num_ste, [])
        has_gagne_recent = any(y >= 2025 and s == 'GAGNE' for y, s in statuts)
        if has_gagne_recent:
            statut_client = 'CLIENT_ACTIF'
        elif any(s == 'GAGNE' for _, s in statuts):
            statut_client = 'CLIENT_ACTIF'
        else:
            statut_client = 'PROSPECT'

        # Prospecteur mapping
        prosp_val = None
        if row['prosp'].upper() in ('ALEX', 'ALEXA'):
            prosp_val = 'ALEX'
        elif row['prosp'].upper() in ('CL',):
            prosp_val = 'CL'

        # Build phone string
        phone = safe_str(row['telephone'])
        if phone == '0':
            phone = ''

        # Build company data
        company_data = {
            'name': row['client_name'],
            'address': {
                'addressStreet1': row['adresse1'],
                'addressStreet2': row.get('adresse2', ''),
                'addressCity': row['ville'],
                'addressPostcode': str(int(float(row['cp']))) if row['cp'] else '',
            },
            'numeroSociete': num_ste,
            'typeClient': type_client,
            'statutClient': statut_client,
        }

        if sous_type:
            company_data['sousType'] = sous_type
        if dept_num:
            company_data['departementNumero'] = dept_num
        if dept_select:
            company_data['departement'] = dept_select
        if prosp_val:
            company_data['prospecteur'] = prosp_val

        # Company has no phones or emails fields in Twenty CRM standard schema
        # Email and phone are on the Person object instead

        result = rest_create('companies', company_data)
        time.sleep(0.7)

        if 'error' in result:
            print(f"  ERR: {row['client_name']} → {json.dumps(result['error'])[:200]}")
            log.append({'type': 'company', 'name': row['client_name'], 'status': 'error', 'error': str(result['error'])[:200]})
            companies_errors += 1
        else:
            company_id = result.get('data', {}).get('createCompany', {}).get('id') or result.get('id')
            if not company_id:
                # Try different response structures
                if 'data' in result:
                    for key in result['data']:
                        if isinstance(result['data'][key], dict) and 'id' in result['data'][key]:
                            company_id = result['data'][key]['id']
                            break

            if company_id:
                mappings['companies'][num_ste] = company_id
                print(f"  OK: {row['client_name']} → {company_id}")
                log.append({'type': 'company', 'name': row['client_name'], 'status': 'ok', 'id': company_id})
                companies_created += 1
            else:
                print(f"  ???: {row['client_name']} → response: {json.dumps(result)[:300]}")
                log.append({'type': 'company', 'name': row['client_name'], 'status': 'unknown', 'response': str(result)[:200]})
                companies_errors += 1

    print(f"  Companies: {companies_created} créées, {companies_errors} erreurs\n")

    # Save mappings
    with open(MAPPINGS_FILE, 'w') as f:
        json.dump(mappings, f, indent=2, ensure_ascii=False)

    # --------------------------------------------------------
    # STEP 4: Import Persons
    # --------------------------------------------------------
    print("4. Import Persons...")
    persons_created = 0
    persons_errors = 0

    for key, row in persons_map.items():
        if key in mappings['persons']:
            print(f"  SKIP: {row['contact']} (déjà importé)")
            continue

        first_name, last_name = parse_contact_name(row['contact'], row['civilite'])
        company_id = mappings['companies'].get(row['num_ste'])

        person_data = {
            'name': {
                'firstName': first_name or '',
                'lastName': last_name or row['contact'],
            },
        }

        if company_id:
            person_data['companyId'] = company_id

        # Email
        if row['email']:
            email = row['email'].split('/')[0].strip().split(' ')[0].strip()
            person_data['emails'] = {'primaryEmail': email}

        # Phone
        phone = safe_str(row['telephone'])
        if phone and phone != '0':
            person_data['phones'] = {
                'primaryPhoneNumber': phone,
                'primaryPhoneCountryCode': 'FR',
                'primaryPhoneCallingCode': '+33'
            }

        result = rest_create('people', person_data)
        time.sleep(0.7)

        if 'error' in result:
            print(f"  ERR: {row['contact']} → {json.dumps(result['error'])[:200]}")
            log.append({'type': 'person', 'name': row['contact'], 'status': 'error', 'error': str(result['error'])[:200]})
            persons_errors += 1
        else:
            person_id = result.get('data', {}).get('createPerson', {}).get('id') or result.get('id')
            if not person_id and 'data' in result:
                for k in result['data']:
                    if isinstance(result['data'][k], dict) and 'id' in result['data'][k]:
                        person_id = result['data'][k]['id']
                        break

            if person_id:
                mappings['persons'][key] = person_id
                print(f"  OK: {row['contact']} → {person_id}")
                log.append({'type': 'person', 'name': row['contact'], 'status': 'ok', 'id': person_id})
                persons_created += 1
            else:
                print(f"  ???: {row['contact']} → {json.dumps(result)[:300]}")
                log.append({'type': 'person', 'name': row['contact'], 'status': 'unknown', 'response': str(result)[:200]})
                persons_errors += 1

    print(f"  Persons: {persons_created} créées, {persons_errors} erreurs\n")

    # Save mappings
    with open(MAPPINGS_FILE, 'w') as f:
        json.dump(mappings, f, indent=2, ensure_ascii=False)

    # --------------------------------------------------------
    # STEP 5: Import Opportunities
    # --------------------------------------------------------
    print("5. Import Opportunities...")
    opps_created = 0
    opps_errors = 0

    for row in all_rows:
        opp_key = row['num_devis'] or f"{row['num_ste']}-{row['year']}-{row['row_idx']}"

        if opp_key in mappings['opportunities']:
            print(f"  SKIP: {opp_key} (déjà importé)")
            continue

        company_id = mappings['companies'].get(row['num_ste'])
        person_key = f"{row['num_ste']}|{row['contact']}"
        person_id = mappings['persons'].get(person_key)

        prestations, nature, modalite = parse_norme(row['norme'])

        # Calculate amount
        amount = None
        if row['offre1']:
            amount = {
                'amountMicros': int(row['offre1'] * 1000000),
                'currencyCode': 'EUR'
            }

        # Calculate remise amount
        montant_remise = None
        taux_remise = None
        if row['offre2'] and row['offre1'] and row['offre1'] > 0:
            montant_remise = {
                'amountMicros': int(row['offre2'] * 1000000),
                'currencyCode': 'EUR'
            }
            taux_remise = round((1 - row['offre2'] / row['offre1']) * 100)

        date_devis = excel_date_to_iso(row['date_devis'])
        date_docs = excel_date_to_iso(row['date_docs'])

        opp_data = {
            'name': opp_key,
            'numeroDevis': row['num_devis'] or opp_key,
            'prestation': prestations,
            'naturePrestation': nature,
            'statutDevis': row['statut'],
            'anneeDevis': row['year'],
            'normeOriginale': row['norme'],
        }

        if amount:
            opp_data['amount'] = amount
        if montant_remise:
            opp_data['montantRemise'] = montant_remise
        if taux_remise is not None:
            opp_data['tauxRemise'] = taux_remise
        if modalite:
            opp_data['modalite'] = modalite
        if date_devis:
            opp_data['dateDevis'] = date_devis
        if date_docs:
            opp_data['dateEnvoiDocs'] = date_docs
        if company_id:
            opp_data['companyId'] = company_id
        if person_id:
            opp_data['pointOfContactId'] = person_id

        result = rest_create('opportunities', opp_data)
        time.sleep(0.7)

        if 'error' in result:
            print(f"  ERR: {opp_key} → {json.dumps(result['error'])[:200]}")
            log.append({'type': 'opportunity', 'name': opp_key, 'status': 'error', 'error': str(result['error'])[:200]})
            opps_errors += 1
        else:
            opp_id = result.get('data', {}).get('createOpportunity', {}).get('id') or result.get('id')
            if not opp_id and 'data' in result:
                for k in result['data']:
                    if isinstance(result['data'][k], dict) and 'id' in result['data'][k]:
                        opp_id = result['data'][k]['id']
                        break

            if opp_id:
                mappings['opportunities'][opp_key] = opp_id
                print(f"  OK: {opp_key} → {opp_id}")
                log.append({'type': 'opportunity', 'name': opp_key, 'status': 'ok', 'id': opp_id})
                opps_created += 1
            else:
                print(f"  ???: {opp_key} → {json.dumps(result)[:300]}")
                log.append({'type': 'opportunity', 'name': opp_key, 'status': 'unknown', 'response': str(result)[:200]})
                opps_errors += 1

    print(f"  Opportunities: {opps_created} créées, {opps_errors} erreurs\n")

    # Save mappings and log
    with open(MAPPINGS_FILE, 'w') as f:
        json.dump(mappings, f, indent=2, ensure_ascii=False)
    with open(LOG_FILE, 'w') as f:
        json.dump(log, f, indent=2, ensure_ascii=False)

    # --------------------------------------------------------
    # SUMMARY
    # --------------------------------------------------------
    print("=" * 60)
    print(f"RÉSUMÉ PHASE 2 (MODE TEST - {TEST_LIMIT} lignes)")
    print(f"  Companies:     {companies_created} créées, {companies_errors} erreurs")
    print(f"  Persons:       {persons_created} créées, {persons_errors} erreurs")
    print(f"  Opportunities: {opps_created} créées, {opps_errors} erreurs")
    print(f"  Mappings sauvés dans: {MAPPINGS_FILE}")
    print(f"  Log sauvé dans: {LOG_FILE}")
    print("=" * 60)

if __name__ == '__main__':
    main()
